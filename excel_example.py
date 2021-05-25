#!/usr/bin/env python
import openpyxl, os

os.chdir('/home/isma/dev/python')

workbook = openpyxl.load_workbook('Untitled_1.xlsx')

sheet1 = workbook[workbook.sheetnames[0]]

cellA1 = sheet1['A1']

print('this is the value of cell A1: ' + str(cellA1.value))
print('the complete column B list is: ')

for i in range(1,8):
	print(i, sheet1.cell(row=i, column=2).value)

#sheet1.title = 'my sheet'

#workbook.create_sheet(index=0, title='first')

sheet0 = workbook[workbook.sheetnames[0]]
sheet0['A1'] = 42

workbook.save('Untitled_1.xlsx')

